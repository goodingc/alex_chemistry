import concurrent.futures
import csv
import re
import threading
from copy import deepcopy, copy
from functools import reduce
from typing import List, Dict, Tuple, Optional

import openpyxl
from openpyxl import Workbook
from rdkit import Chem

from src.Models import Complex, Molecule
from src.MoleculeValueFinder import MoleculeValueFinder
from src.utils import remove_titanium, remove_bridge, get_ligands, get_bridge_idty, \
    replace_rounds, display_table

LigandResult = Tuple[List[Optional[str]], Optional[str]]


def metallocene(smiles: str) -> LigandResult:
    molecule = reduce(
        lambda mol, smiles:
        Chem.DeleteSubstructs(
            mol,
            Chem.MolFromSmiles(smiles, sanitize=False)
        ),
        [
            'Cl[Ti]Cl',
            '[Ti]',
            '[Cl-]',
        ],
        smiles
    )
    root_pattern = Chem.MolFromSmiles("c1cccc1.c1cccc1", sanitize=False)
    chains = Chem.ReplaceCore(molecule, root_pattern, labelByIndex=True)
    pieces = Chem.GetMolFrags(chains, asMols=True)
    ligands = [Chem.MolToSmiles(x, True) for x in pieces]
    rs = []
    for ligand in ligands:
        ligand = replace_rounds(ligand, [(r'\[C+@+H*\]', 'C'), (r"\[\d\*\]", "*"), (r'\[N+@+H*\]', 'N')])
        if ligand.count("*") > 1:
            rs.append(ligand.split('C(*)')[0])
            rs.append("*C" + ligand.split("C(*)")[1])
            return [*rs, *([None] * (6 - len(rs)))], None
        if len(ligand) < 5:
            continue
        else:
            rs.append(re.sub(r"\[\d\*\]", "*", ligand))
    ligands = [*rs, *([""] * (2 - len(rs)))]
    return [*ligands, *([None] * (6 - len(ligands)))], None


def onno(molecule: str) -> LigandResult:
    molecule = remove_titanium(molecule)
    molecule_br = remove_bridge(molecule, "NCCN", [1, 2])
    bridge_identity = get_bridge_idty(molecule, "N.N")
    if len(bridge_identity) > 1:
        return [bridge_identity[1], *get_ligands(molecule_br, "NCC1=CC=CC=C1O", [3, 4, 5, 6, 1])], bridge_identity[0]
    return get_ligands(molecule_br, "NCC1=CC=CC=C1O", [0, 3, 4, 5, 6, 1]), bridge_identity[0]


def ono(molecule: str) -> LigandResult:
    molecule = remove_titanium(molecule)
    return get_ligands(molecule, "NCC1=CC=CC=C1O", [0, 3, 4, 5, 6, 1]), None


def onnoen(molecule: str) -> LigandResult:
    molecule = remove_titanium(molecule)
    molecule_br = remove_bridge(molecule, "NC(C=CC=C1)=C1N", [1, 6])
    return [None, *get_ligands(molecule_br, "OC1=CC=CC=C1C=N", [5, 4, 3, 2, 0])], get_bridge_idty(molecule, "N.N")[0]


def onnoalen(molecule: str) -> LigandResult:
    molecule = remove_titanium(molecule)
    try_1 = remove_bridge(molecule, "NCCN", [1, 2])
    molecule_br = remove_bridge(molecule, "NC(C=CC=C1)=C1N", [1, 6]) if try_1 is None else try_1
    try_1 = get_ligands(molecule_br, "NCC1=CC=CC=C1O", [0, 3, 4, 5, 6, 1])
    if try_1 is None:
        return [None, *get_ligands(molecule_br, "OC1=CC=CC=C1C=N", [5, 4, 3, 2, 7])], get_bridge_idty(molecule, "N.N")[
            0]
    return try_1, get_bridge_idty(molecule, "N.N")[0]


def onoen(molecule: str) -> LigandResult:
    molecule = remove_titanium(molecule)
    return get_ligands(molecule, "OC1=CC=CC=C1C=N", [8, 5, 4, 3, 2, 7]), None


extraction_dictionary = {
    'M': metallocene,
    'ONNO': onno,
    'ONO': ono,
    'ONNOen': onnoen,
    'ONNOalen': onnoalen,
    'ONOen': onoen
}


class ComplexProcessor:
    complexes: Dict[int, Complex]

    def __init__(self, input_workbook_path, limit=-1):
        workbook = openpyxl.load_workbook(input_workbook_path, data_only=True)
        input_sheet = workbook['ChemOffice1']

        self.complexes = {}

        row = 2
        name = input_sheet.cell(row=row, column=1).value
        while name is not None:
            name = input_sheet.cell(row=row, column=1).value
            class_name = input_sheet.cell(row=row, column=3).value
            if class_name not in extraction_dictionary:
                row += 1
                continue
            smiles = input_sheet.cell(row=row, column=2).value
            smiles = replace_rounds(smiles, [
                (r'\[Del\]', ''),
                (r'\(\)', ''),
            ])
            self.complexes[row] = Complex(smiles, class_name)
            if row == limit:
                break
            row += 1

    def extract_ligands(self):
        for _, complex in self.complexes.items():
            ligand_smiles, bridge_smiles = extraction_dictionary[complex.class_name](Chem.MolFromSmiles(
                complex.smiles,
                replacements={
                    '[Del]': '',
                    '()': '',
                },
                sanitize=False
            ))
            complex.ligands = list(map(lambda l: None if l is None else Molecule(l), ligand_smiles))
            complex.bridge = None if bridge_smiles is None else Molecule(bridge_smiles)

    def find_ligand_values_thread_pool(self, data_file_path: str):
        value_finder = MoleculeValueFinder(data_file_path)

        def find_values(data: Tuple[Complex, MoleculeValueFinder]):
            complex = data[0]
            value_finder = data[1]
            for ligand in complex.ligands:
                if ligand is not None:
                    ligand.set_values(value_finder)
            if complex.bridge is not None:
                complex.bridge.set_values(value_finder)
            return complex

        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            executor.map(find_values, map(lambda c: (c, copy(value_finder)), self.complexes.values()))
            # value_finder_count = 5
            # value_finders = [copy(value_finder) for _ in range(value_finder_count)]
            # futures = []
            # for index, complex in enumerate(self.complexes.values()):
            #     futures.append(executor.submit(find_values, (complex, value_finders[index % value_finder_count])))
            # concurrent.futures.wait(futures)

    def find_ligand_values(self, data_file_path: str):
        value_finder = MoleculeValueFinder(data_file_path)
        for complex in self.complexes.values():
            for ligand in complex.ligands:
                if ligand is not None:
                    ligand.set_values(value_finder)
            if complex.bridge is not None:
                complex.bridge.set_values(value_finder)

    def display(self):
        display_table([
            'Row',
            'Complex',
            'Class',
            'R<sup>1</sup>',
            'R<sup>2</sup>',
            'R<sup>3</sup>',
            'R<sup>4</sup>',
            'R<sup>5</sup>',
            'R<sup>6</sup>',
            'Bridge'
        ], list(map(lambda row: [row, *self.complexes[row].get_row_elements()], self.complexes)))

    def write_to_sheet(self, output_workbook_path: str):
        workbook = Workbook()
        sheet = workbook.active

        titles = [
            'Complex',
            'Class',
            'R1',
            'HanschPi1',
            'Hammett1',
            'VdW1',
            'R2',
            'HanschPi2',
            'Hammett2',
            'VdW2',
            'R3',
            'HanschPi3',
            'Hammett3',
            'VdW3',
            'R4',
            'HanschPi4',
            'Hammett4',
            'VdW4',
            'R5',
            'HanschPi5',
            'Hammett5',
            'VdW5',
            'R6',
            'HanschPi6',
            'Hammett6',
            'VdW6',
            'Bridge',
            'HanschPi7',
            'Hammett7',
            'VdW7',
        ]

        for (column, value) in zip(range(len(titles)), titles):
            sheet.cell(1, column + 1, value)

        with open('./res/csvDescriptors.csv', 'w', newline='') as csvdescriptors:
            wr = csv.writer(csvdescriptors, quoting=csv.QUOTE_ALL)
            wr.writerow(titles)
            for row in self.complexes:
                complex = [*self.complexes[row].get_xcl_row_elements()]
                sheet.cell(row, 1, complex[0].smiles)
                sheet.cell(row, 2, complex[1])
                row_to_write = [complex[0].smiles, complex[1]]
                for i in range(7):
                    ligand = complex[i + 2]
                    if type(ligand) is not str:
                        values = [ligand.smiles, ligand.hansch_pi, ligand.hammett, ligand.vd_w_volume]
                    else:
                        values = [ligand, 'N/A', 'N/A', 'N/A']
                    for n in range(4):
                        sheet.cell(row, ((4 * (i + 1)) + n) - 1,
                                   values[n])
                        row_to_write.append(values[n])
                wr.writerow(row_to_write)

        workbook.save(output_workbook_path)

    def read_from_csv(self, path: str):
        with open(path, 'r', newline='') as file:
            reader = csv.reader(file)
            for row in reader:
                pass
